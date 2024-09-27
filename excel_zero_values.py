import openpyxl
from openpyxl import load_workbook


def sostituisci_celle_vuote(file_excel):
    try:
        # Apre il file Excel
        wb = load_workbook(file_excel)

        # Cicla su ogni foglio nel file Excel
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Cicla su ogni riga e colonna nel foglio
            for row in sheet.iter_rows():
                for cell in row:
                    # Controlla se la cella è vuota e non è una cella unita
                    if cell.value is None and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        cell.value = 0

        # Salva le modifiche
        wb.save(file_excel)
        print(f"Modifiche salvate con successo nel file {file_excel}")

    except Exception as e:
        print(f"Si è verificato un errore: {e}")


# Esempio di utilizzo
if __name__ == "__main__":
    file_excel = "Teams/Old/Padova_2023_2024.xlsx"  # Inserisci il percorso del tuo file Excel
    sostituisci_celle_vuote(file_excel)
